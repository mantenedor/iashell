#!/usr/bin/env python3

import csv

import hashlib

import json

import mimetypes

import os

import re

import shutil

from datetime import datetime

from pathlib import Path

from typing import List, Dict, Optional



from dotenv import load_dotenv

from openpyxl import load_workbook

from PIL import Image





def load_local_env():

    env_file = Path(os.environ.get("IA_ENV_FILE", "conf/.env"))

    if env_file.exists():

        load_dotenv(env_file)





load_local_env()



# Configuração de diretórios

BASE_DIR = Path("/opt/ia/data/iashell/base")

KNOWLEDGE_DIR = BASE_DIR / "knowledge"

KNOWLEDGE_RAW_DIR = BASE_DIR / "knowledge" / ".raw"

PARSED_DIR = BASE_DIR / "parsed"

TMP_DIR = Path(os.environ.get("TMP_DIR", "/opt/ia/data/iashell/base/tmp"))

TMP_RAW_DIR = Path(os.environ.get("TMP_RAW_DIR", "/opt/ia/data/iashell/base/tmp/.raw"))

DOCS_INDEX_FILE = BASE_DIR / "document_catalog.json"

CHUNKS_FILE = BASE_DIR / "chunks.jsonl"





def ensure_dir(path: Path):

    path.mkdir(parents=True, exist_ok=True)





def ensure_parent(path: Path):

    path.parent.mkdir(parents=True, exist_ok=True)





def ensure_runtime_structure():

    """Garante que toda a estrutura de diretórios existe"""

    for d in [TMP_DIR, TMP_RAW_DIR, KNOWLEDGE_DIR, KNOWLEDGE_RAW_DIR, PARSED_DIR]:

        ensure_dir(d)



    ensure_parent(DOCS_INDEX_FILE)

    ensure_parent(CHUNKS_FILE)



    if not DOCS_INDEX_FILE.exists():

        DOCS_INDEX_FILE.write_text("[]\n", encoding="utf-8")



    if not CHUNKS_FILE.exists():

        CHUNKS_FILE.write_text("", encoding="utf-8")





def sha256_file(path: Path):

    h = hashlib.sha256()

    with open(path, "rb") as f:

        for chunk in iter(lambda: f.read(1024 * 1024), b""):

            h.update(chunk)

    return h.hexdigest()





def file_size(path: Path):

    return path.stat().st_size





def now_iso():

    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"





def slugify(value):

    value = value.lower().strip()

    value = re.sub(r"[^\w\s.-]", "-", value)

    value = re.sub(r"\s+", "-", value)

    return value.strip("-._")





def load_index():

    ensure_runtime_structure()

    with open(DOCS_INDEX_FILE, "r", encoding="utf-8") as f:

        return json.load(f)





def save_index(index_data):

    ensure_parent(DOCS_INDEX_FILE)

    with open(DOCS_INDEX_FILE, "w", encoding="utf-8") as f:

        json.dump(index_data, f, ensure_ascii=False, indent=2)





def list_documents():

    return load_index()





def get_doc_metadata(doc_id: str) -> dict:

    """Retorna metadados de um documento específico"""

    docs = load_index()

    for doc in docs:

        if doc.get("doc_id") == doc_id:

            return doc

    return None





def rewrite_chunks_for_doc(doc_id: str, new_chunks: list):

    ensure_runtime_structure()



    remaining_chunks = []



    if CHUNKS_FILE.exists():

        with open(CHUNKS_FILE, "r", encoding="utf-8") as f:

            for line in f:

                line = line.strip()

                if not line:

                    continue

                obj = json.loads(line)

                if obj.get("doc_id") != doc_id:

                    remaining_chunks.append(obj)



    remaining_chunks.extend(new_chunks)



    ensure_parent(CHUNKS_FILE)



    with open(CHUNKS_FILE, "w", encoding="utf-8") as f:

        for item in remaining_chunks:

            f.write(json.dumps(item, ensure_ascii=False) + "\n")





def detect_doc_type(path: Path):

    ext = path.suffix.lower()



    # Documentos de texto puro

    if ext in {".txt", ".md", ".log", ".json", ".yaml", ".yml", ".ini", ".cfg"}:

        return "text"

    

    # HTML

    if ext in {".htm", ".html"}:

        return "html"



    # CSV/TSV

    if ext in {".csv", ".tsv"}:

        return "csv"



    # Planilhas Excel

    if ext in {".xlsx", ".xlsm"}:

        return "spreadsheet"



    # PDF

    if ext in {".pdf"}:

        return "pdf"



    # Imagens

    if ext in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}:

        return "image"



    # Fallback para MIME type

    mime, _ = mimetypes.guess_type(str(path))

    if mime and mime.startswith("text/"):

        return "text"



    return "binary"





def make_doc_id(path: Path, file_hash: str):

    return f"{slugify(path.stem)}-{file_hash[:12]}"





def read_text_file(path: Path):

    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]

    for enc in encodings:

        try:

            return path.read_text(encoding=enc)

        except UnicodeDecodeError:

            continue

    return path.read_text(encoding="utf-8", errors="replace")





def summarize_text(text: str, limit: int = 400) -> str:

    if not text:

        return ""

    text = re.sub(r"\s+", " ", text).strip()

    return text[:limit].rstrip() + ("..." if len(text) > limit else "")





# ========== FUNÇÕES PARA HTML ==========



def extract_text_from_html_with_bs4(file_path: str) -> str:

    """Extrai texto de HTML usando BeautifulSoup e html2text"""

    try:

        from bs4 import BeautifulSoup

        import html2text

        

        # Lê o conteúdo do arquivo

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:

            html_content = f.read()

        

        # Parseia com BeautifulSoup

        soup = BeautifulSoup(html_content, 'html.parser')

        

        # Remove elementos que não contribuem para o conteúdo

        for elem in soup(["script", "style", "nav", "footer", "header", "aside"]):

            elem.decompose()

        

        # Configura o html2text

        h2t = html2text.HTML2Text()

        h2t.ignore_links = False

        h2t.ignore_images = True

        h2t.ignore_emphasis = False

        h2t.body_width = 0  # Não quebrar linhas

        

        # Converte para Markdown

        markdown_text = h2t.handle(str(soup))

        

        return markdown_text

        

    except ImportError:

        print("Aviso: beautifulsoup4 ou html2text não instalados. Usando fallback.")

        return extract_text_from_html_fallback(file_path)

    except Exception as e:

        return f"[Erro ao processar HTML: {e}]"





def extract_text_from_html_fallback(file_path: str) -> str:

    """Fallback simples removendo tags HTML"""

    try:

        import re

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:

            html = f.read()

        # Remove tags HTML

        text = re.sub(r'<[^>]+>', ' ', html)

        text = re.sub(r'\s+', ' ', text).strip()

        return text

    except Exception as e:

        return f"[Erro no fallback: {e}]"





def parse_html(path: Path):

    """Parseia arquivo HTML e retorna estrutura de seções"""

    try:

        from bs4 import BeautifulSoup

        

        with open(path, 'r', encoding='utf-8', errors='ignore') as f:

            html_content = f.read()

        

        soup = BeautifulSoup(html_content, 'html.parser')

        

        # Remove elementos irrelevantes

        for elem in soup(["script", "style", "nav", "footer", "header", "aside"]):

            elem.decompose()

        

        # Título da página

        title_tag = soup.find('title')

        title = title_tag.get_text().strip() if title_tag else path.stem

        

        # Metadados úteis

        meta_description = ""

        meta_desc = soup.find('meta', attrs={'name': 'description'})

        if meta_desc and meta_desc.get('content'):

            meta_description = meta_desc['content']

        

        # Divide em seções baseadas em cabeçalhos

        sections = []

        current_section = {"title": "Introdução", "content": [], "level": 1}

        

        # Processa o corpo do documento

        body = soup.body if soup.body else soup

        

        for elem in body.children:

            if not hasattr(elem, 'name') or not elem.name:

                continue

                

            if elem.name and elem.name.startswith('h') and len(elem.name) == 2:

                # Finaliza seção anterior

                if current_section["content"]:

                    sections.append(current_section)

                

                # Nova seção

                level = int(elem.name[1])

                current_section = {

                    "title": elem.get_text().strip(),

                    "content": [],

                    "level": level

                }

            else:

                # Adiciona conteúdo à seção atual

                text = elem.get_text().strip()

                if text:

                    current_section["content"].append(text)

        

        # Adiciona última seção

        if current_section["content"]:

            sections.append(current_section)

        

        # Se não encontrou seções, cria uma única seção com todo o texto

        if not sections:

            full_text = body.get_text().strip()

            sections = [{"title": "Conteúdo", "content": [full_text], "level": 1}]

        

        return {

            "doc_id": path.stem,

            "title": title,

            "meta_description": meta_description,

            "sections": sections

        }

        

    except ImportError:

        # Fallback sem BeautifulSoup

        text = extract_text_from_html_fallback(str(path))

        return {

            "doc_id": path.stem,

            "title": path.stem,

            "meta_description": "",

            "sections": [{"title": "Conteúdo", "content": [text], "level": 1}]

        }

    except Exception as e:

        return {

            "doc_id": path.stem,

            "title": path.stem,

            "meta_description": "",

            "sections": [{"title": "Erro", "content": [f"Erro ao parsear HTML: {e}"], "level": 1}]

        }





# ========== FUNÇÕES PARA PDF ==========



def extract_text_from_pdf_with_pdfplumber(pdf_path: str) -> str:

    """Extrai texto de PDF usando pdfplumber (mais robusto)"""

    try:

        import pdfplumber

        text_parts = []

        

        with pdfplumber.open(pdf_path) as pdf:

            for page_num, page in enumerate(pdf.pages, 1):

                try:

                    # Tenta extrair texto

                    text = page.extract_text() or ""

                    

                    # Se não conseguir texto, tenta extrair tabelas

                    if not text.strip():

                        tables = page.extract_tables()

                        if tables:

                            table_text = []

                            for table in tables:

                                for row in table:

                                    if row:

                                        table_text.append(" | ".join([str(cell) if cell else "" for cell in row]))

                            text = "\n".join(table_text)

                    

                    # Se ainda não tiver texto, tenta extrair palavras individuais

                    if not text.strip():

                        words = page.extract_words()

                        if words:

                            text = " ".join([word.get('text', '') for word in words])

                    

                    # Última tentativa: extract_text com layout

                    if not text.strip():

                        text = page.extract_text(layout=True) or ""

                    

                    if text and text.strip():

                        text_parts.append(f"--- PÁGINA {page_num} ---\n{text}")

                    else:

                        text_parts.append(f"--- PÁGINA {page_num} (sem texto extraível) ---")

                            

                except Exception as e:

                    text_parts.append(f"--- PÁGINA {page_num} (erro: {e}) ---")

        

        if not text_parts:

            return "[PDF sem conteúdo extraído]"

        

        return "\n\n".join(text_parts)

        

    except ImportError:

        print("pdfplumber não instalado. Usando PyPDF2 como fallback.")

        return extract_text_from_pdf_fallback(pdf_path)

    except Exception as e:

        return f"[Erro ao processar PDF com pdfplumber: {e}]"





def extract_text_from_pdf_fallback(pdf_path: str) -> str:

    """Fallback usando PyPDF2"""

    try:

        from PyPDF2 import PdfReader

        text_parts = []

        

        with open(pdf_path, 'rb') as file:

            reader = PdfReader(file)

            

            for page_num, page in enumerate(reader.pages, 1):

                try:

                    text = page.extract_text() or ""

                    

                    if text and text.strip():

                        text_parts.append(f"--- PÁGINA {page_num} ---\n{text}")

                    else:

                        text_parts.append(f"--- PÁGINA {page_num} (sem texto extraível) ---")

                        

                except Exception as e:

                    text_parts.append(f"--- PÁGINA {page_num} (erro: {e}) ---")

        

        if not text_parts:

            return "[PDF sem conteúdo extraído]"

        

        return "\n\n".join(text_parts)

        

    except ImportError:

        return "[PyPDF2 não instalado]"

    except Exception as e:

        return f"[Erro no fallback: {e}]"





def extract_text_from_pdf(pdf_path: str) -> str:

    """Wrapper principal para extração de texto de PDF"""

    try:

        import pdfplumber

        return extract_text_from_pdf_with_pdfplumber(pdf_path)

    except ImportError:

        return extract_text_from_pdf_fallback(pdf_path)





def parse_pdf_with_pdfplumber(path: Path):

    """Parseia PDF usando pdfplumber"""

    try:

        import pdfplumber

        pages = []

        

        with pdfplumber.open(str(path)) as pdf:

            for page_num, page in enumerate(pdf.pages, start=1):

                try:

                    text = page.extract_text() or ""

                    

                    if not text.strip():

                        tables = page.extract_tables()

                        if tables:

                            table_texts = []

                            for table in tables:

                                for row in table:

                                    if row:

                                        table_texts.append(" | ".join([str(cell) if cell else "" for cell in row]))

                            text = "\n".join(table_texts)

                    

                    if not text.strip():

                        words = page.extract_words()

                        if words:

                            text = " ".join([word.get('text', '') for word in words])

                    

                    if not text.strip():

                        text = page.extract_text(layout=True) or ""

                    

                    pages.append({"page": page_num, "text": text or ""})

                    

                except Exception as e:

                    pages.append({"page": page_num, "text": f"[Erro: {e}]"})

        

        return pages

        

    except ImportError:

        return parse_pdf_fallback(path)





def parse_pdf_fallback(path: Path):

    """Fallback usando PyPDF2"""

    try:

        from PyPDF2 import PdfReader

        reader = PdfReader(str(path))

        pages = []

        

        for page_num, page in enumerate(reader.pages, start=1):

            try:

                text = page.extract_text() or ""

                pages.append({"page": page_num, "text": text})

            except Exception as e:

                pages.append({"page": page_num, "text": f"[Erro: {e}]"})

        

        return pages

        

    except Exception as e:

        return [{"page": 1, "text": f"[Erro ao processar PDF: {e}]"}]





def parse_pdf(path: Path):

    """Wrapper principal para parse de PDF"""

    try:

        import pdfplumber

        return parse_pdf_with_pdfplumber(path)

    except ImportError:

        return parse_pdf_fallback(path)





# ========== FUNÇÕES PARA PLANILHAS ==========



def parse_spreadsheet(path: Path):

    wb = load_workbook(path, data_only=True, read_only=True)

    sheets = []



    for ws in wb.worksheets:

        rows = []

        row_limit = 300

        col_limit = 30



        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):

            if row_idx > row_limit:

                break

            normalized = []

            for cell in list(row)[:col_limit]:

                normalized.append("" if cell is None else str(cell))

            rows.append(normalized)



        preview_lines = []

        for row in rows[:80]:

            if any(x.strip() for x in row):

                preview_lines.append(" | ".join(row))



        sheets.append({

            "name": ws.title,

            "rows": rows,

            "preview_text": "\n".join(preview_lines)

        })



    wb.close()

    return sheets





# ========== FUNÇÕES PARA CSV ==========



def parse_csv(path: Path):

    ext = path.suffix.lower()

    delimiter = "\t" if ext == ".tsv" else ","



    rows = []

    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:

        reader = csv.reader(f, delimiter=delimiter)

        for row in reader:

            rows.append(row)



    preview_text = "\n".join([" | ".join(map(str, row)) for row in rows[:200]])

    return {

        "name": path.stem,

        "rows": rows[:2000],

        "preview_text": preview_text

    }





# ========== FUNÇÕES PARA IMAGENS ==========



def parse_image(path: Path):

    width = None

    height = None

    mode = None

    try:

        with Image.open(path) as img:

            width, height = img.size

            mode = img.mode

    except Exception:

        pass



    sidecar_txt = path.with_suffix(path.suffix + ".txt")

    sidecar_md = path.with_suffix(path.suffix + ".md")



    description = ""

    if sidecar_txt.exists():

        description = read_text_file(sidecar_txt).strip()

    elif sidecar_md.exists():

        description = read_text_file(sidecar_md).strip()



    return {

        "width": width,

        "height": height,

        "mode": mode,

        "description": description,

    }





# ========== FUNÇÃO PRINCIPAL DE PARSING ==========



def parse_document(meta):

    """Parseia documento baseado no tipo"""

    raw_path = Path(meta["path_raw"])

    doc_type = meta["type"]



    # Texto puro

    if doc_type == "text":

        return {

            "doc_id": meta["doc_id"],

            "type": "text",

            "title": raw_path.name,

            "sections": [{"section": "body", "text": read_text_file(raw_path)}],

        }

    

    # HTML

    if doc_type == "html":

        parsed = parse_html(raw_path)

        return {

            "doc_id": meta["doc_id"],

            "type": "html",

            "title": parsed["title"],

            "meta_description": parsed.get("meta_description", ""),

            "sections": parsed["sections"],

        }



    # PDF

    if doc_type == "pdf":

        return {

            "doc_id": meta["doc_id"],

            "type": "pdf",

            "title": raw_path.name,

            "pages": parse_pdf(raw_path),

        }



    # CSV

    if doc_type == "csv":

        return {

            "doc_id": meta["doc_id"],

            "type": "csv",

            "title": raw_path.name,

            "tables": [parse_csv(raw_path)],

        }



    # Planilhas

    if doc_type == "spreadsheet":

        return {

            "doc_id": meta["doc_id"],

            "type": "spreadsheet",

            "title": raw_path.name,

            "sheets": parse_spreadsheet(raw_path),

        }



    # Imagens

    if doc_type == "image":

        image_meta = parse_image(raw_path)

        return {

            "doc_id": meta["doc_id"],

            "type": "image",

            "title": raw_path.name,

            "image_meta": {

                "width": image_meta["width"],

                "height": image_meta["height"],

                "mode": image_meta["mode"],

            },

            "description": image_meta["description"],

        }



    # Tipo não suportado

    return {

        "doc_id": meta["doc_id"],

        "type": doc_type,

        "title": raw_path.name,

        "unsupported": True,

    }





def save_parsed_document(parsed_doc: dict, parsed_path: Path):

    ensure_parent(parsed_path)

    with open(parsed_path, "w", encoding="utf-8") as f:

        json.dump(parsed_doc, f, ensure_ascii=False, indent=2)





def normalize_whitespace(text: str) -> str:

    if not text:

        return ""

    text = text.replace("\r", "")

    text = re.sub(r"\n{3,}", "\n\n", text)

    text = re.sub(r"[ \t]{2,}", " ", text)

    return text.strip()





def split_chunks(text, size=1800, overlap=250):

    """Divide texto em chunks de forma inteligente"""

    if not text or not text.strip():

        return []



    text = normalize_whitespace(text)

    

    if len(text) < size:

        return [text]

    

    chunks = []

    start = 0

    text_len = len(text)

    

    while start < text_len:

        end = min(start + size, text_len)

        

        if end < text_len:

            search_start = max(start, end - 200)

            search_text = text[search_start:end]

            

            cut_points = []

            

            for pattern, priority in [('. ', 10), ('.\n', 10), ('\n\n', 9), ('\n', 8), ('.', 7), (' ', 1)]:

                last_pos = search_text.rfind(pattern)

                if last_pos != -1:

                    cut_points.append((search_start + last_pos + len(pattern), priority))

            

            if cut_points:

                cut_points.sort(key=lambda x: x[1], reverse=True)

                best_cut = cut_points[0][0]

                if best_cut > start:

                    end = best_cut

        

        chunk = text[start:end].strip()

        if chunk:

            chunks.append(chunk)

        

        if end >= text_len:

            break

            

        start = max(end - overlap, start + 1)

    

    return chunks if chunks else [text[:size]]





def extract_keywords(text: str):

    """Extrai palavras-chave relevantes do texto"""

    words = re.findall(r"[a-zA-Z0-9_./:-]{3,}", text.lower())

    stop = {

        "the", "and", "for", "with", "that", "this", "from", "como", "para",

        "uma", "com", "sem", "dos", "das", "que", "por", "não", "ser", "são",

        "json", "text", "page", "sheet", "pdf", "file", "document", "arquivo",

        "página", "capítulo", "seção", "figura", "tabela", "html", "body", "div"

    }

    keywords = []

    seen = set()

    for w in words:

        if w in stop or len(w) < 3:

            continue

        if w not in seen:

            seen.add(w)

            keywords.append(w)

    return keywords[:80]





def chunk_document(parsed_doc):

    """Gera chunks a partir do documento parseado"""

    doc_id = parsed_doc["doc_id"]

    doc_type = parsed_doc["type"]

    title = parsed_doc.get("title", doc_id)

    chunks = []

    

    def add_chunk(text: str, source_ref: str, extra=None):

        if not text or not text.strip():

            return

        extra = extra or {}

        for idx, part in enumerate(split_chunks(text), start=1):

            if part and part.strip():

                chunk = {

                    "doc_id": doc_id,

                    "chunk_id": f"{doc_id}::{source_ref}::{idx:03d}",

                    "title": title,

                    "type": doc_type,

                    "source_ref": source_ref,

                    "text": part,

                    "keywords": extract_keywords(title + "\n" + part),

                }

                chunk.update(extra)

                chunks.append(chunk)



    # Texto puro

    if doc_type == "text":

        for section in parsed_doc.get("sections", []):

            text = section.get("text", "")

            if text:

                add_chunk(text, f"section:{section.get('section', 'body')}")



    # HTML

    elif doc_type == "html":

        sections = parsed_doc.get("sections", [])

        for section in sections:

            section_title = section.get("title", "Conteúdo")

            content = "\n".join(section.get("content", []))

            if content:

                add_chunk(content, f"section:{section_title}", {"level": section.get("level", 1)})

        

        # Adiciona meta description se existir

        meta_desc = parsed_doc.get("meta_description", "")

        if meta_desc and meta_desc.strip():

            add_chunk(f"Meta Description: {meta_desc}", "metadata")



    # PDF

    elif doc_type == "pdf":

        pages = parsed_doc.get("pages", [])

        if pages:

            for page in pages:

                text = page.get("text", "")

                if text and text.strip() and not text.startswith("[Erro") and not text.startswith("--- PÁGINA"):

                    add_chunk(text, f"page:{page.get('page', 1)}", {"page": page.get("page", 1)})

        

        if not chunks:

            try:

                meta = get_doc_metadata(doc_id)

                if meta:

                    raw_path = Path(meta["path_raw"])

                    if raw_path.exists():

                        text = extract_text_from_pdf(str(raw_path))

                        if text and text.strip() and not text.startswith("[Erro"):

                            add_chunk(text, "full_document")

            except Exception as e:

                print(f"Erro no fallback PDF: {e}")



    # CSV

    elif doc_type == "csv":

        for table in parsed_doc.get("tables", []):

            text = table.get("preview_text", "")

            if text:

                add_chunk(text, f"table:{table.get('name', 'default')}")



    # Planilhas

    elif doc_type == "spreadsheet":

        for sheet in parsed_doc.get("sheets", []):

            text = sheet.get("preview_text", "")

            if text:

                add_chunk(text, f"sheet:{sheet.get('name', 'Sheet1')}")



    # Imagens

    elif doc_type == "image":

        text = f"Imagem: {title}\nDescrição: {parsed_doc.get('description', '')}\nMetadados: {parsed_doc.get('image_meta', {})}"

        add_chunk(text, "image:description")



    # Se não gerou chunks, cria pelo menos um chunk com metadados

    if not chunks:

        add_chunk(f"Documento: {title}\nTipo: {doc_type}\nSem conteúdo extraível", "metadata")

    

    return chunks





def build_summary_from_parsed(parsed_doc: dict) -> str:

    """Gera resumo do documento de forma robusta"""

    doc_type = parsed_doc.get("type")

    text_parts = []

    

    try:

        # HTML

        if doc_type == "html":

            sections = parsed_doc.get("sections", [])

            for section in sections[:3]:

                content = "\n".join(section.get("content", [])[:5])

                if content:

                    text_parts.append(f"{section.get('title', 'Seção')}: {content[:200]}")

            

            meta_desc = parsed_doc.get("meta_description", "")

            if meta_desc:

                text_parts.insert(0, f"Descrição: {meta_desc}")



        # PDF

        elif doc_type == "pdf":

            pages = parsed_doc.get("pages", [])

            for page in pages[:3]:

                text = page.get("text", "").strip()

                if text and not text.startswith("[Erro") and not text.startswith("--- PÁGINA"):

                    lines = text.split('\n')[:10]

                    text_parts.extend(lines)



        # Texto puro

        elif doc_type == "text":

            sections = parsed_doc.get("sections", [])

            for section in sections[:2]:

                text = section.get("text", "").strip()

                if text:

                    lines = text.split('\n')[:15]

                    text_parts.extend(lines)



        # CSV

        elif doc_type == "csv":

            tables = parsed_doc.get("tables", [])

            for table in tables[:1]:

                text = table.get("preview_text", "").strip()

                if text:

                    text_parts.append(f"CSV Preview:\n{text[:500]}")



        # Planilhas

        elif doc_type == "spreadsheet":

            sheets = parsed_doc.get("sheets", [])

            for sheet in sheets[:2]:

                text = sheet.get("preview_text", "").strip()

                if text:

                    text_parts.append(f"Planilha '{sheet.get('name', '')}':\n{text[:300]}")



        # Imagens

        elif doc_type == "image":

            desc = parsed_doc.get("description", "").strip()

            meta = parsed_doc.get("image_meta", {})

            if desc:

                text_parts.append(desc)

            if meta:

                text_parts.append(f"Dimensões: {meta.get('width', '?')}x{meta.get('height', '?')}")

        

        if text_parts:

            full_text = " ".join(text_parts)

            full_text = re.sub(r'\s+', ' ', full_text)

            return summarize_text(full_text, limit=600)

        

    except Exception as e:

        print(f"Erro ao gerar resumo: {e}")

    

    return f"Documento {parsed_doc.get('title', '')} - resumo não disponível"





def pick_raw_dir(source_mode: str) -> Path:

    if source_mode == "tmp":

        return TMP_RAW_DIR

    return KNOWLEDGE_RAW_DIR





def ingest_file(path: str, source_mode="manual"):

    ensure_runtime_structure()



    src = Path(path)

    if not src.exists() or not src.is_file():

        raise FileNotFoundError(f"Arquivo não encontrado: {src}")



    file_hash = sha256_file(src)

    doc_id = make_doc_id(src, file_hash)

    doc_type = detect_doc_type(src)



    raw_root = pick_raw_dir(source_mode)

    raw_filename = f"{doc_id}{src.suffix.lower()}"

    raw_path = raw_root / raw_filename

    parsed_path = PARSED_DIR / f"{doc_id}.json"



    ensure_dir(raw_root)



    index_data = load_index()

    existing = next((x for x in index_data if x.get("doc_id") == doc_id), None)

    if existing:

        return existing



    if not raw_path.exists():

        shutil.copy2(src, raw_path)



    meta = {

        "doc_id": doc_id,

        "filename_original": src.name,

        "filename_raw": raw_filename,

        "path_source": str(src.resolve()),

        "path_raw": str(raw_path),

        "path_parsed": str(parsed_path),

        "type": doc_type,

        "hash_sha256": file_hash,

        "size_bytes": file_size(src),

        "created_at": now_iso(),

        "status": "ingested",

        "source_mode": source_mode,

        "summary": "",

        "chunk_count": 0,

    }



    index_data.append(meta)

    save_index(index_data)

    return meta





def reindex_document(doc_id):

    """Reindexa documento: parseia, gera chunks e resumo"""

    index_data = load_index()

    meta = next(x for x in index_data if x["doc_id"] == doc_id)



    # Parseia o documento

    parsed_doc = parse_document(meta)

    save_parsed_document(parsed_doc, Path(meta["path_parsed"]))



    # Gera chunks

    chunks = chunk_document(parsed_doc)

    

    if chunks:

        rewrite_chunks_for_doc(doc_id, chunks)

        meta["chunk_count"] = len(chunks)

    else:

        print(f"Aviso: Nenhum chunk gerado para {doc_id}")

        meta["chunk_count"] = 0



    meta["parsed_at"] = now_iso()

    meta["status"] = "parsed"

    

    # Gera resumo

    summary = build_summary_from_parsed(parsed_doc)

    if not summary or summary == f"Documento {parsed_doc.get('title', '')} - resumo não disponível":

        if chunks and len(chunks) > 0:

            chunk_text = chunks[0].get("text", "")

            if chunk_text:

                summary = summarize_text(chunk_text[:1000], limit=400)

        else:

            summary = f"Documento {meta.get('filename_original', 'sem nome')} - {meta.get('type', 'desconhecido')}"

    

    meta["summary"] = summary



    save_index(index_data)

    return meta





def ingest_and_index(path, source_mode="manual"):

    """Ingere e indexa um documento"""

    meta = ingest_file(path, source_mode)

    return reindex_document(meta["doc_id"])





def upsert_document(path: str, source_mode: str):

    """Atualiza ou insere documento"""

    ensure_runtime_structure()



    src = Path(path)

    if not src.exists() or not src.is_file():

        return None



    current_hash = sha256_file(src)

    index_data = load_index()



    existing = next(

        (x for x in index_data if x.get("path_source") == str(src.resolve())),

        None

    )



    if existing and existing.get("hash_sha256") == current_hash:

        return None



    if existing:

        old_doc_id = existing["doc_id"]

        index_data = [x for x in index_data if x.get("doc_id") != old_doc_id]

        save_index(index_data)

        rewrite_chunks_for_doc(old_doc_id, [])



        old_parsed = Path(existing.get("path_parsed", ""))

        if old_parsed.exists():

            old_parsed.unlink()



    return ingest_and_index(str(src), source_mode)





def sync_directory(source_dir: Path, source_mode: str):

    """Sincroniza diretório com a base de conhecimento"""

    ensure_runtime_structure()

    changed = []



    for item in sorted(source_dir.iterdir()):

        if item.name.startswith("."):

            continue

        if not item.is_file():

            continue



        result = upsert_document(str(item), source_mode)

        if result is not None:

            changed.append(result)



    return changed





def sync_all_sources():

    """Sincroniza todas as fontes de documentos"""

    changed = []

    changed.extend(sync_directory(TMP_DIR, "tmp"))

    changed.extend(sync_directory(KNOWLEDGE_DIR, "knowledge"))

    return changed





def load_all_chunks():

    """Carrega todos os chunks do arquivo"""

    ensure_runtime_structure()

    chunks = []



    if CHUNKS_FILE.exists():

        with open(CHUNKS_FILE, "r", encoding="utf-8") as f:

            for line in f:

                line = line.strip()

                if line:

                    try:

                        chunks.append(json.loads(line))

                    except json.JSONDecodeError:

                        continue



    return chunks





def get_document_chunks(doc_id: str):

    """Retorna chunks de um documento específico"""

    return [c for c in load_all_chunks() if c.get("doc_id") == doc_id]





def build_doc_context(doc_id: str, limit: int = 8) -> str:

    """Constrói contexto a partir dos chunks do documento"""

    chunks = get_document_chunks(doc_id)

    

    if chunks:

        parts = []

        for i, chunk in enumerate(chunks[:limit]):

            if chunk.get('source_ref') == 'metadata' and any(c.get('source_ref') != 'metadata' for c in chunks):

                continue

                

            source = chunk.get('source_ref', f'parte {i+1}')

            text = chunk.get('text', '')

            if text and text.strip() and not text.startswith('[Erro') and 'sem conteúdo extraível' not in text:

                truncated = text[:1500] + "..." if len(text) > 1500 else text

                parts.append(f"[{source}]\n{truncated}")

        

        if parts:

            return "\n\n".join(parts)

    

    # Fallback para arquivo parsed

    try:

        meta = get_doc_metadata(doc_id)

        if meta and meta.get("path_parsed"):

            parsed_path = Path(meta["path_parsed"])

            if parsed_path.exists():

                with open(parsed_path, 'r', encoding='utf-8') as f:

                    parsed = json.load(f)

                

                if parsed.get("type") == "html":

                    sections = parsed.get("sections", [])

                    text_parts = []

                    for section in sections[:3]:

                        title = section.get('title', 'Seção')

                        content = "\n".join(section.get('content', [])[:5])

                        if content:

                            text_parts.append(f"{title}:\n{content[:300]}")

                    if text_parts:

                        return "Conteúdo HTML:\n\n" + "\n\n".join(text_parts)

                

                elif parsed.get("type") == "pdf":

                    pages = parsed.get("pages", [])

                    text_parts = []

                    for page in pages[:3]:

                        page_text = page.get("text", "")

                        if page_text and page_text.strip() and not page_text.startswith("[Erro"):

                            text_parts.append(f"Página {page.get('page', '?')}:\n{page_text[:300]}")

                    if text_parts:

                        return "Conteúdo do PDF:\n\n" + "\n\n".join(text_parts)

                

                elif parsed.get("type") == "text":

                    sections = parsed.get("sections", [])

                    if sections:

                        text = sections[0].get("text", "")

                        return text[:3000] if text else ""

    except Exception as e:

        print(f"Erro ao ler documento: {e}")

    

    return ""





def format_document_catalog(limit: int = 50) -> str:

    """Formata catálogo de documentos para exibição"""

    docs = load_index()

    docs = sorted(docs, key=lambda x: x.get("created_at", ""), reverse=True)[:limit]



    if not docs:

        return ""



    parts = []

    for d in docs:

        summary = d.get('summary', '')

        if summary and len(summary) > 100:

            summary = summary[:97] + "..."

        

        chunk_info = f"chunks: {d.get('chunk_count', 0)}"

        if d.get('chunk_count', 0) == 1 and 'sem conteúdo' in summary.lower():

            chunk_info = "⚠️ Sem conteúdo extraível"

        

        parts.append(

            "\n".join([

                f"doc_id: {d.get('doc_id', '')}",

                f"nome: {d.get('filename_original', '')}",

                f"tipo: {d.get('type', '')}",

                f"origem: {d.get('source_mode', '')}",

                f"caminho: {d.get('path_source', '')}",

                f"tamanho_bytes: {d.get('size_bytes', 0)}",

                f"chunks: {chunk_info}",

                f"resumo: {summary}",

            ])

        )



    return "\n\n".join(parts)





def search_chunks(query, limit=5):

    """Busca chunks por relevância"""

    query_terms = set(re.findall(r"[a-zA-Z0-9_./:-]{2,}", query.lower()))

    if not query_terms:

        return []



    results = []

    for chunk in load_all_chunks():

        if chunk.get('source_ref') == 'metadata' and 'sem conteúdo' in chunk.get('text', '').lower():

            continue

            

        haystack = (

            chunk.get("title", "").lower() + "\n" +

            chunk.get("text", "").lower() + "\n" +

            " ".join(chunk.get("keywords", []))

        )



        score = 0

        for term in query_terms:

            if term in haystack:

                score += 5

            if term in chunk.get("title", "").lower():

                score += 3

            if term in " ".join(chunk.get("keywords", [])):

                score += 2



        if score > 0:

            results.append((score, chunk))



    results.sort(key=lambda x: x[0], reverse=True)

    return [item[1] for item in results[:limit]]





def build_context_block(query, limit=5):

    """Constrói bloco de contexto baseado em busca"""

    hits = search_chunks(query, limit)



    parts = []

    for h in hits:

        label = f"{h.get('doc_id', '')} | {h.get('source_ref', '')}"

        text = h.get('text', '')

        if text and not text.startswith('[Erro') and 'sem conteúdo' not in text.lower():

            truncated = text[:800] + "..." if len(text) > 800 else text

            parts.append(f"[{label}]\n{truncated}")



    return "\n\n".join(parts)





def diagnose_document(doc_id: str):

    """Função de diagnóstico para verificar processamento do documento"""

    print(f"\n{'='*60}")

    print(f"DIAGNÓSTICO DO DOCUMENTO: {doc_id}")

    print(f"{'='*60}\n")

    

    docs = load_index()

    doc = next((d for d in docs if d["doc_id"] == doc_id), None)

    if not doc:

        print("❌ ERRO: Documento não encontrado no catálogo!")

        return

    

    print("📋 METADADOS:")

    print(f"  Nome: {doc.get('filename_original')}")

    print(f"  Tipo: {doc.get('type')}")

    print(f"  Tamanho: {doc.get('size_bytes')} bytes")

    print(f"  Chunks: {doc.get('chunk_count', 0)}")

    print(f"  Status: {doc.get('status')}")

    print(f"  Resumo: {doc.get('summary', '')[:200]}")

    

    raw_path = Path(doc.get("path_raw", ""))

    if raw_path.exists():

        print(f"\n📄 ARQUIVO RAW: {raw_path}")

        print(f"  Tamanho: {raw_path.stat().st_size} bytes")

    else:

        print(f"\n❌ ARQUIVO RAW NÃO ENCONTRADO: {raw_path}")

    

    parsed_path = Path(doc.get("path_parsed", ""))

    if parsed_path.exists():

        print(f"\n📄 ARQUIVO PARSED: {parsed_path}")

        print(f"  Tamanho: {parsed_path.stat().st_size} bytes")

        try:

            with open(parsed_path, 'r', encoding='utf-8') as f:

                parsed = json.load(f)

                

                if parsed.get("type") == "html":

                    sections = parsed.get("sections", [])

                    print(f"  Seções: {len(sections)}")

                    if sections:

                        for section in sections[:2]:

                            title = section.get('title', 'Sem título')

                            content = "\n".join(section.get('content', [])[:2])

                            if content:

                                print(f"\n  Seção '{title}':")

                                print(f"  {content[:200]}...")

                

                elif parsed.get("type") == "pdf":

                    pages = parsed.get("pages", [])

                    print(f"  Páginas: {len(pages)}")

                    if pages:

                        for page in pages[:2]:

                            page_text = page.get('text', '')

                            if page_text and not page_text.startswith('--- PÁGINA') and not page_text.startswith('[Erro'):

                                print(f"\n  Página {page.get('page', '?')}:")

                                print(f"  {page_text[:200]}...")

                                break

                

                elif parsed.get("type") == "text":

                    sections = parsed.get("sections", [])

                    if sections:

                        text = sections[0].get('text', '')

                        print(f"\n  Conteúdo (início):")

                        print(f"  {text[:200]}...")

                        

        except Exception as e:

            print(f"  ❌ Erro ao ler parsed: {e}")

    else:

        print(f"\n❌ ARQUIVO PARSED NÃO ENCONTRADO: {parsed_path}")

    

    chunks = get_document_chunks(doc_id)

    print(f"\n🔍 CHUNKS: {len(chunks)} encontrados")

    if chunks:

        chunks_com_conteudo = [c for c in chunks if c.get('text') and not c.get('text').startswith('[Erro') and 'sem conteúdo' not in c.get('text', '').lower()]

        print(f"  Chunks com conteúdo real: {len(chunks_com_conteudo)}")

        

        for i, chunk in enumerate(chunks[:3]):

            print(f"\n  Chunk {i+1}:")

            print(f"    ID: {chunk.get('chunk_id', 'N/A')}")

            print(f"    Source: {chunk.get('source_ref', 'N/A')}")

            text = chunk.get('text', '')

            if len(text) > 150:

                text = text[:150] + "..."

            print(f"    Texto: {text}")

    else:

        print("  ⚠️ NENHUM chunk encontrado!")

    

    print(f"\n{'='*60}\n")

    return doc
